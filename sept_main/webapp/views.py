from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.views import View
from django.contrib.auth.models import User
from django.contrib.auth.models import Group, Permission
from requests.exceptions import ConnectionError
import requests
from django.views.decorators.csrf import csrf_exempt
import json
from django.http import JsonResponse
from .models import *

from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import UploadExcelForm
from openpyxl import load_workbook
from .models import Product
import os
from django.contrib import messages
from openpyxl import load_workbook
import os

def upload_excel(request):
    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES)
        error_message = None

        if form.is_valid():
            excel_file = request.FILES['excel_file']
            if not excel_file:
                error_message = "Please select a file to upload."
            else:
                file_extension = os.path.splitext(excel_file.name)[1].lower()
                if file_extension != '.xlsx':
                    error_message = "Only Excel files (.xlsx) are allowed for upload."
                    
                else:
                    wb = load_workbook(excel_file)
                    ws = wb.active
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        name, price, description = row
                        Product.objects.create(name=name, price=price, description=description)
                    messages.success(request, 'Excel data imported successfully!')

                    return redirect('upload_excel')
        else:
            error_message = "Form validation failed. Please check your inputs."

        
        messages.error(request, error_message)

    else:
        form = UploadExcelForm()

    return render(request, 'upload_excel.html', {'form': form})


def test(request):
    rrr = movie.objects.get(movie_name="RRR")
    characters_with_movies = rrr.character.all()
    messages.success(request, 'Excel data imported successfully!')
    return render(request, 'register.html',{"error":"error"})


@csrf_exempt
def login_user(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))
            username = data.get('username')
            password = data.get('password')
            response_data = { 
                'message': 'Data received successfully',
                'username': username,
                'password': password,
            }

            return JsonResponse(response_data)

        except Exception as e:
            response_data = {'error': str(e)}
            return JsonResponse(response_data, status=400)
    return JsonResponse({'message': 'This view only accepts POST requests.'}, status=405)


def register(request):
    user=User.objects.filter(username="bhuvan").values()
    
    return render(request,'register.html',{"user":user})

from .forms import RegistrationForm
from django.shortcuts import render, redirect
from django.contrib.auth import login

def register1(request):
    user = request.user
    is_member = user.groups.filter(name='manger').exists()
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('profile')  
    else:
        form = RegistrationForm()
    return render(request, 'register.html', {'form': form,'is_member':is_member})